import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

url=input("URL을 입력하세요: ")

headers={
    'user-Agent':'Mozilla/5.0',
    'Content-Type':'text/html; charset=utf-8'
}

response=requests.get(url,headers=headers)

results=re.findall(r'[\w\.-]+@[\w\.-]+',response.text)
results=list(set(results))

print(results)

try:
    wb=load_workbook(r".\email.xlsx",data_only=True)
    sheet=wb.active
except:
    wb=Workbook()
    sheet=wb.active

for result in results:
    sheet.append([result])

wb.save(r"./email.xlsx")